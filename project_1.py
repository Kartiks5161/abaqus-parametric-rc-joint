"""
Project 1: Automated 2D RC Frame Analysis & Preliminary Design

This program generates and analyses a 2D multi-storey reinforced-concrete
moment frame using the matrix stiffness method. It produces:

- undeformed/deformed frame plots
- beam shear-force and bending-moment diagrams
- nodal displacements and support reactions
- preliminary IS-code-inspired member checks
- PDF and Excel summary reports

The design checks are intentionally preliminary and educational. A full
building design must be reviewed using the complete latest IS provisions,
load combinations, detailing rules, and project-specific engineering judgement.
"""

from __future__ import annotations

import argparse
import datetime as dt
import math
import os
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

MISSING_DEPENDENCY = None
try:
    import numpy as np
    from PIL import Image, ImageDraw
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image as PdfImage,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlsxImage
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ModuleNotFoundError as exc:
    MISSING_DEPENDENCY = exc.name


CONCRETE_DENSITY_KN_M3 = 25.0
IS1893_ZONE_FACTORS = {"II": 0.10, "III": 0.16, "IV": 0.24, "V": 0.36}


@dataclass(frozen=True)
class LoadCombination:
    name: str
    dead_factor: float
    live_factor: float
    eq_factor: float
    purpose: str


SERVICE_COMBINATION = LoadCombination("Service: DL + LL + EQ", 1.0, 1.0, 1.0, "service")
DESIGN_COMBINATIONS = [
    LoadCombination("1.5(DL + LL)", 1.5, 1.5, 0.0, "strength"),
    LoadCombination("1.2(DL + LL + EQ)", 1.2, 1.2, 1.2, "strength"),
    LoadCombination("1.2(DL + LL - EQ)", 1.2, 1.2, -1.2, "strength"),
    LoadCombination("1.5(DL + EQ)", 1.5, 0.0, 1.5, "strength"),
    LoadCombination("1.5(DL - EQ)", 1.5, 0.0, -1.5, "strength"),
    LoadCombination("0.9DL + 1.5EQ", 0.9, 0.0, 1.5, "strength"),
    LoadCombination("0.9DL - 1.5EQ", 0.9, 0.0, -1.5, "strength"),
]


def ensure_dependencies() -> None:
    if MISSING_DEPENDENCY is None:
        return
    print("\nA required Python package is missing:", MISSING_DEPENDENCY)
    print("\nInstall the project requirements first:")
    print("  python -m pip install -r requirements.txt")
    print("\nThen run the program again:")
    print("  python project_1.py --gui")
    raise SystemExit(1)


@dataclass
class FrameInput:
    storeys: int = 3
    bays: int = 2
    bay_width_m: float = 4.0
    storey_height_m: float = 3.0
    beam_width_mm: float = 230.0
    beam_depth_mm: float = 450.0
    column_width_mm: float = 300.0
    column_depth_mm: float = 500.0
    concrete_grade_mpa: float = 25.0
    steel_grade_mpa: float = 500.0
    dead_load_kn_m: float = 12.0
    live_load_kn_m: float = 6.0
    seismic_zone: str = "III"
    lateral_load_kn: float = 0.0
    importance_factor: float = 1.0
    response_reduction_factor: float = 5.0
    support_condition: str = "fixed"


@dataclass
class Node:
    id: int
    x: float
    y: float


@dataclass
class Element:
    id: int
    node_i: int
    node_j: int
    kind: str
    area_m2: float
    inertia_m4: float
    udl_kn_m: float = 0.0


@dataclass
class AnalysisResult:
    nodes: List[Node]
    elements: List[Element]
    displacements: np.ndarray
    reactions: np.ndarray
    global_loads: np.ndarray
    member_end_forces: Dict[int, np.ndarray]
    beam_diagrams: Dict[int, Dict[str, List[float]]]
    checks: Dict[int, Dict[str, str]]
    verification: Dict[str, Tuple[str, str]]
    load_combination: LoadCombination
    design_envelope: Dict[str, object]
    benchmark_results: List[Dict[str, str]]
    base_shear_kn: float
    storey_forces_kn: List[float]
    inputs: FrameInput


def ask_int(prompt: str, default: int) -> int:
    value = input(f"{prompt} [{default}]: ").strip()
    return default if not value else int(value)


def ask_float(prompt: str, default: float) -> float:
    value = input(f"{prompt} [{default}]: ").strip()
    return default if not value else float(value)


def ask_text(prompt: str, default: str) -> str:
    value = input(f"{prompt} [{default}]: ").strip()
    return default if not value else value


def read_inputs_interactively() -> FrameInput:
    print("\nAutomated 2D RC Frame Analysis & Preliminary Design")
    print("Press Enter to accept the default value shown in brackets.\n")
    data = FrameInput()
    data.storeys = ask_int("Number of storeys", data.storeys)
    data.bays = ask_int("Number of bays", data.bays)
    data.bay_width_m = ask_float("Bay width in m", data.bay_width_m)
    data.storey_height_m = ask_float("Storey height in m", data.storey_height_m)
    data.beam_width_mm = ask_float("Beam width in mm", data.beam_width_mm)
    data.beam_depth_mm = ask_float("Beam depth in mm", data.beam_depth_mm)
    data.column_width_mm = ask_float("Column width in mm", data.column_width_mm)
    data.column_depth_mm = ask_float("Column depth in mm", data.column_depth_mm)
    data.concrete_grade_mpa = ask_float("Concrete grade fck in MPa", data.concrete_grade_mpa)
    data.steel_grade_mpa = ask_float("Steel grade fy in MPa", data.steel_grade_mpa)
    data.dead_load_kn_m = ask_float("Dead load on beams in kN/m", data.dead_load_kn_m)
    data.live_load_kn_m = ask_float("Live load on beams in kN/m", data.live_load_kn_m)
    data.seismic_zone = ask_text("Seismic zone: II, III, IV, or V", data.seismic_zone).upper()
    data.lateral_load_kn = ask_float(
        "Optional total lateral load in kN, enter 0 to calculate from seismic zone",
        data.lateral_load_kn,
    )
    data.importance_factor = ask_float("Importance factor", data.importance_factor)
    data.response_reduction_factor = ask_float(
        "Response reduction factor", data.response_reduction_factor
    )
    data.support_condition = ask_text(
        "Support condition: fixed, pinned, or roller", data.support_condition
    ).lower()
    validate_inputs(data)
    return data


def validate_inputs(data: FrameInput) -> None:
    if data.storeys < 1 or data.bays < 1:
        raise ValueError("Storeys and bays must be at least 1.")
    if data.bay_width_m <= 0 or data.storey_height_m <= 0:
        raise ValueError("Bay width and storey height must be positive.")
    if min(
        data.beam_width_mm,
        data.beam_depth_mm,
        data.column_width_mm,
        data.column_depth_mm,
        data.concrete_grade_mpa,
        data.steel_grade_mpa,
    ) <= 0:
        raise ValueError("Material grades and member dimensions must be positive.")
    if data.seismic_zone not in IS1893_ZONE_FACTORS:
        raise ValueError("Seismic zone must be II, III, IV, or V.")
    if data.support_condition not in {"fixed", "pinned", "roller"}:
        raise ValueError("Support condition must be fixed, pinned, or roller.")


def section_properties(width_mm: float, depth_mm: float) -> Tuple[float, float]:
    b = width_mm / 1000.0
    d = depth_mm / 1000.0
    area = b * d
    inertia = b * d**3 / 12.0
    return area, inertia


def generate_frame(data: FrameInput) -> Tuple[List[Node], List[Element]]:
    nodes: List[Node] = []
    for level in range(data.storeys + 1):
        for grid in range(data.bays + 1):
            nodes.append(
                Node(
                    id=level * (data.bays + 1) + grid,
                    x=grid * data.bay_width_m,
                    y=level * data.storey_height_m,
                )
            )

    beam_area, beam_i = section_properties(data.beam_width_mm, data.beam_depth_mm)
    col_area, col_i = section_properties(data.column_width_mm, data.column_depth_mm)

    elements: List[Element] = []
    eid = 0
    for storey in range(data.storeys):
        for grid in range(data.bays + 1):
            n1 = storey * (data.bays + 1) + grid
            n2 = (storey + 1) * (data.bays + 1) + grid
            elements.append(Element(eid, n1, n2, "column", col_area, col_i))
            eid += 1

    for level in range(1, data.storeys + 1):
        for bay in range(data.bays):
            n1 = level * (data.bays + 1) + bay
            n2 = n1 + 1
            elements.append(Element(eid, n1, n2, "beam", beam_area, beam_i))
            eid += 1

    return nodes, elements


def beam_self_weight_kn_m(data: FrameInput) -> float:
    beam_area, _ = section_properties(data.beam_width_mm, data.beam_depth_mm)
    return beam_area * CONCRETE_DENSITY_KN_M3


def beam_udl_for_combination(data: FrameInput, combination: LoadCombination) -> float:
    dead_with_self_weight = data.dead_load_kn_m + beam_self_weight_kn_m(data)
    return combination.dead_factor * dead_with_self_weight + combination.live_factor * data.live_load_kn_m


def element_geometry(nodes: List[Node], element: Element) -> Tuple[float, float, float]:
    ni = nodes[element.node_i]
    nj = nodes[element.node_j]
    dx = nj.x - ni.x
    dy = nj.y - ni.y
    length = math.hypot(dx, dy)
    return length, dx / length, dy / length


def transformation(c: float, s: float) -> np.ndarray:
    return np.array(
        [
            [c, s, 0, 0, 0, 0],
            [-s, c, 0, 0, 0, 0],
            [0, 0, 1, 0, 0, 0],
            [0, 0, 0, c, s, 0],
            [0, 0, 0, -s, c, 0],
            [0, 0, 0, 0, 0, 1],
        ],
        dtype=float,
    )


def local_stiffness(e_modulus_kn_m2: float, area_m2: float, inertia_m4: float, length_m: float) -> np.ndarray:
    ea_l = e_modulus_kn_m2 * area_m2 / length_m
    ei = e_modulus_kn_m2 * inertia_m4
    l = length_m
    return np.array(
        [
            [ea_l, 0, 0, -ea_l, 0, 0],
            [0, 12 * ei / l**3, 6 * ei / l**2, 0, -12 * ei / l**3, 6 * ei / l**2],
            [0, 6 * ei / l**2, 4 * ei / l, 0, -6 * ei / l**2, 2 * ei / l],
            [-ea_l, 0, 0, ea_l, 0, 0],
            [0, -12 * ei / l**3, -6 * ei / l**2, 0, 12 * ei / l**3, -6 * ei / l**2],
            [0, 6 * ei / l**2, 2 * ei / l, 0, -6 * ei / l**2, 4 * ei / l],
        ],
        dtype=float,
    )


def dof_indices(element: Element) -> List[int]:
    return [
        3 * element.node_i,
        3 * element.node_i + 1,
        3 * element.node_i + 2,
        3 * element.node_j,
        3 * element.node_j + 1,
        3 * element.node_j + 2,
    ]


def fixed_end_load_local(element: Element, length_m: float) -> np.ndarray:
    if element.kind != "beam" or element.udl_kn_m == 0:
        return np.zeros(6)
    q = -element.udl_kn_m
    return np.array([0, q * length_m / 2, q * length_m**2 / 12, 0, q * length_m / 2, -q * length_m**2 / 12])


def spectral_acceleration_medium_soil(period_s: float) -> float:
    if period_s <= 0.10:
        return 1.0 + 15.0 * period_s
    if period_s <= 0.55:
        return 2.5
    if period_s <= 4.0:
        return 1.36 / period_s
    return 0.34 / period_s**2


def calculate_lateral_loads(data: FrameInput) -> Tuple[float, List[float]]:
    total_width = data.bays * data.bay_width_m
    beam_area, _ = section_properties(data.beam_width_mm, data.beam_depth_mm)
    beam_self = beam_area * CONCRETE_DENSITY_KN_M3
    live_fraction = 0.25 if data.live_load_kn_m <= 3.0 else 0.50
    floor_weight = (data.dead_load_kn_m + beam_self + live_fraction * data.live_load_kn_m) * total_width
    seismic_weight = floor_weight * data.storeys

    if data.lateral_load_kn > 0:
        base_shear = data.lateral_load_kn
    else:
        height = data.storeys * data.storey_height_m
        period = 0.075 * height**0.75
        z = IS1893_ZONE_FACTORS[data.seismic_zone]
        sa_g = spectral_acceleration_medium_soil(period)
        ah = (z / 2.0) * (data.importance_factor / data.response_reduction_factor) * sa_g
        base_shear = ah * seismic_weight

    heights = np.array([(i + 1) * data.storey_height_m for i in range(data.storeys)])
    weights = np.full(data.storeys, floor_weight)
    distribution = weights * heights**2
    forces = base_shear * distribution / distribution.sum()
    return float(base_shear), [float(force) for force in forces]


def restrained_dofs(data: FrameInput) -> List[int]:
    restrained: List[int] = []
    for grid in range(data.bays + 1):
        base = 3 * grid
        if data.support_condition == "fixed":
            restrained.extend([base, base + 1, base + 2])
        elif data.support_condition == "pinned":
            restrained.extend([base, base + 1])
        else:
            restrained.append(base + 1)
    return restrained


def analyse_frame(
    data: FrameInput,
    combination: LoadCombination = SERVICE_COMBINATION,
    include_design_envelope: bool = True,
) -> AnalysisResult:
    validate_inputs(data)
    nodes, elements = generate_frame(data)
    beam_udl = beam_udl_for_combination(data, combination)
    for element in elements:
        if element.kind == "beam":
            element.udl_kn_m = beam_udl
    ndof = len(nodes) * 3
    stiffness = np.zeros((ndof, ndof))
    loads = np.zeros(ndof)
    e_modulus = 5000.0 * math.sqrt(data.concrete_grade_mpa) * 1000.0

    element_cache: Dict[int, Tuple[np.ndarray, np.ndarray, np.ndarray, float]] = {}
    for element in elements:
        length, c, s = element_geometry(nodes, element)
        k_local = local_stiffness(e_modulus, element.area_m2, element.inertia_m4, length)
        t = transformation(c, s)
        k_global = t.T @ k_local @ t
        fe_local = fixed_end_load_local(element, length)
        fe_global = t.T @ fe_local
        indices = dof_indices(element)
        for a, ia in enumerate(indices):
            loads[ia] += fe_global[a]
            for b, ib in enumerate(indices):
                stiffness[ia, ib] += k_global[a, b]
        element_cache[element.id] = (k_local, t, fe_local, length)

    unfactored_base_shear, unfactored_storey_forces = calculate_lateral_loads(data)
    base_shear = unfactored_base_shear * combination.eq_factor
    storey_forces = [force * combination.eq_factor for force in unfactored_storey_forces]
    for storey_index, force in enumerate(storey_forces, start=1):
        floor_nodes = [storey_index * (data.bays + 1) + grid for grid in range(data.bays + 1)]
        for node_id in floor_nodes:
            loads[3 * node_id] += force / len(floor_nodes)

    restrained = restrained_dofs(data)
    free = [dof for dof in range(ndof) if dof not in restrained]
    if not free:
        raise ValueError("No free degrees of freedom are available for analysis.")

    displacements = np.zeros(ndof)
    k_ff = stiffness[np.ix_(free, free)]
    f_f = loads[free]
    displacements[free] = np.linalg.solve(k_ff, f_f)
    reactions = stiffness @ displacements - loads

    member_end_forces: Dict[int, np.ndarray] = {}
    beam_diagrams: Dict[int, Dict[str, List[float]]] = {}
    for element in elements:
        k_local, t, fe_local, length = element_cache[element.id]
        u_global = displacements[dof_indices(element)]
        u_local = t @ u_global
        end_force = k_local @ u_local - fe_local
        member_end_forces[element.id] = end_force
        if element.kind == "beam":
            beam_diagrams[element.id] = beam_sfd_bmd(end_force, element.udl_kn_m, length)

    design_envelope = build_design_envelope(data) if include_design_envelope else {}
    checks = preliminary_checks(data, elements, member_end_forces, beam_diagrams, design_envelope)
    benchmark_results = run_benchmark_checks(data) if include_design_envelope else []
    verification = verification_checks(data, nodes, stiffness, displacements, reactions, loads, restrained, base_shear, storey_forces)
    return AnalysisResult(
        nodes=nodes,
        elements=elements,
        displacements=displacements,
        reactions=reactions,
        global_loads=loads,
        member_end_forces=member_end_forces,
        beam_diagrams=beam_diagrams,
        checks=checks,
        verification=verification,
        load_combination=combination,
        design_envelope=design_envelope,
        benchmark_results=benchmark_results,
        base_shear_kn=base_shear,
        storey_forces_kn=storey_forces,
        inputs=data,
    )


def verification_checks(
    data: FrameInput,
    nodes: List[Node],
    stiffness: np.ndarray,
    displacements: np.ndarray,
    reactions: np.ndarray,
    loads: np.ndarray,
    restrained: List[int],
    base_shear: float,
    storey_forces: List[float],
) -> Dict[str, Tuple[str, str]]:
    support_reactions = np.zeros_like(reactions)
    support_reactions[restrained] = reactions[restrained]
    residual = support_reactions + loads
    applied_h = float(sum(loads[0::3]))
    reaction_h = float(sum(support_reactions[0::3]))
    horizontal_residual = sum(residual[0::3])
    applied_v = float(sum(loads[1::3]))
    reaction_v = float(sum(support_reactions[1::3]))
    vertical_residual = sum(residual[1::3])
    applied_moment = 0.0
    reaction_moment = 0.0
    moment_residual = 0.0
    for node in nodes:
        load_fx = loads[3 * node.id]
        load_fy = loads[3 * node.id + 1]
        load_mz = loads[3 * node.id + 2]
        reaction_fx = support_reactions[3 * node.id]
        reaction_fy = support_reactions[3 * node.id + 1]
        reaction_mz = support_reactions[3 * node.id + 2]
        applied_moment += load_mz + node.x * load_fy - node.y * load_fx
        reaction_moment += reaction_mz + node.x * reaction_fy - node.y * reaction_fx
        fx = residual[3 * node.id]
        fy = residual[3 * node.id + 1]
        mz = residual[3 * node.id + 2]
        moment_residual += mz + node.x * fy - node.y * fx
    free_dofs = [dof for dof in range(len(loads)) if dof not in restrained]
    free_residual = stiffness @ displacements - loads
    max_free_residual = max(abs(free_residual[dof]) for dof in free_dofs)
    base_shear_distribution_error = abs(sum(storey_forces) - base_shear)
    horizontal_percent = abs(horizontal_residual) / max(abs(applied_h), 1e-9) * 100.0
    vertical_percent = abs(vertical_residual) / max(abs(applied_v), 1e-9) * 100.0
    moment_percent = abs(moment_residual) / max(abs(applied_moment), 1e-9) * 100.0

    def force_summary(applied: float, reaction: float, residual_value: float, percent: float, unit: str) -> str:
        return f"Applied={applied:.3f} {unit}, Reaction={reaction:.3f} {unit}, Residual={residual_value:.3e} {unit}, Error={percent:.3e}%"

    return {
        "Horizontal force equilibrium": (
            force_summary(applied_h, reaction_h, horizontal_residual, horizontal_percent, "kN"),
            "Applied lateral loads plus support reactions should balance.",
        ),
        "Vertical force equilibrium": (
            force_summary(applied_v, reaction_v, vertical_residual, vertical_percent, "kN"),
            "Gravity loads plus support reactions should balance.",
        ),
        "Global moment equilibrium about origin": (
            f"Applied={applied_moment:.3f} kNm, Reaction={reaction_moment:.3f} kNm, Residual={moment_residual:.3e} kNm, Error={moment_percent:.3e}%",
            "Applied-load moments plus reaction moments should balance.",
        ),
        "Free-joint stiffness equation residual": (
            f"Max |K*u - F| at free DOF = {max_free_residual:.3e}",
            "Small numerical residual means the matrix equations were solved consistently.",
        ),
        "Base shear distribution": (
            f"Base shear={base_shear:.3f} kN, Sum of storey forces={sum(storey_forces):.3f} kN, Difference={base_shear_distribution_error:.3e} kN",
            "Storey lateral forces should add back to the base shear.",
        ),
    }


def beam_sfd_bmd(end_force: np.ndarray, udl_kn_m: float, length_m: float) -> Dict[str, List[float]]:
    xs = np.linspace(0, length_m, 25)
    left_shear = end_force[1]
    left_moment = end_force[2]
    shear = left_shear - udl_kn_m * xs
    moment = left_moment + left_shear * xs - udl_kn_m * xs**2 / 2.0
    return {
        "x": xs.tolist(),
        "shear": shear.tolist(),
        "moment": moment.tolist(),
        "max_shear": float(max(abs(v) for v in shear)),
        "max_moment": float(max(abs(v) for v in moment)),
    }


def build_design_envelope(data: FrameInput) -> Dict[str, object]:
    member_envelope: Dict[int, Dict[str, object]] = {}
    combination_rows: List[Dict[str, str]] = []
    summary = {
        "max_beam_moment": 0.0,
        "max_beam_moment_combo": "",
        "max_beam_shear": 0.0,
        "max_beam_shear_combo": "",
        "max_column_axial": 0.0,
        "max_column_axial_combo": "",
        "max_lateral_displacement": 0.0,
        "max_lateral_displacement_combo": "",
    }

    for combination in DESIGN_COMBINATIONS:
        combo_result = analyse_frame(data, combination, include_design_envelope=False)
        max_sway = max(abs(combo_result.displacements[3 * node.id]) for node in combo_result.nodes) * 1000
        max_beam_moment = max((diagram["max_moment"] for diagram in combo_result.beam_diagrams.values()), default=0.0)
        max_beam_shear = max((diagram["max_shear"] for diagram in combo_result.beam_diagrams.values()), default=0.0)
        max_column_axial = 0.0

        for element in combo_result.elements:
            forces = combo_result.member_end_forces[element.id]
            envelope = member_envelope.setdefault(
                element.id,
                {
                    "max_moment": 0.0,
                    "moment_combo": "",
                    "max_shear": 0.0,
                    "shear_combo": "",
                    "max_axial": 0.0,
                    "axial_combo": "",
                },
            )
            if element.kind == "beam":
                diagram = combo_result.beam_diagrams[element.id]
                if diagram["max_moment"] > envelope["max_moment"]:
                    envelope["max_moment"] = diagram["max_moment"]
                    envelope["moment_combo"] = combination.name
                if diagram["max_shear"] > envelope["max_shear"]:
                    envelope["max_shear"] = diagram["max_shear"]
                    envelope["shear_combo"] = combination.name
            else:
                axial = max(abs(forces[0]), abs(forces[3]))
                max_column_axial = max(max_column_axial, axial)
                if axial > envelope["max_axial"]:
                    envelope["max_axial"] = axial
                    envelope["axial_combo"] = combination.name

        combination_rows.append(
            {
                "Combination": combination.name,
                "Max sway (mm)": f"{max_sway:.3f}",
                "Max beam moment (kNm)": f"{max_beam_moment:.2f}",
                "Max beam shear (kN)": f"{max_beam_shear:.2f}",
                "Max column axial (kN)": f"{max_column_axial:.2f}",
            }
        )
        if max_sway > summary["max_lateral_displacement"]:
            summary["max_lateral_displacement"] = max_sway
            summary["max_lateral_displacement_combo"] = combination.name
        if max_beam_moment > summary["max_beam_moment"]:
            summary["max_beam_moment"] = max_beam_moment
            summary["max_beam_moment_combo"] = combination.name
        if max_beam_shear > summary["max_beam_shear"]:
            summary["max_beam_shear"] = max_beam_shear
            summary["max_beam_shear_combo"] = combination.name
        if max_column_axial > summary["max_column_axial"]:
            summary["max_column_axial"] = max_column_axial
            summary["max_column_axial_combo"] = combination.name

    return {"members": member_envelope, "combinations": combination_rows, "summary": summary}


def required_beam_steel_mm2(moment_knm: float, data: FrameInput) -> float:
    effective_depth = max(data.beam_depth_mm - 40.0, data.beam_depth_mm * 0.85)
    lever_arm = 0.87 * effective_depth
    return moment_knm * 1_000_000.0 / max(0.87 * data.steel_grade_mpa * lever_arm, 1e-9)


def minimum_beam_steel_mm2(data: FrameInput) -> float:
    effective_depth = max(data.beam_depth_mm - 40.0, data.beam_depth_mm * 0.85)
    return 0.85 * data.beam_width_mm * effective_depth / data.steel_grade_mpa


def nominal_stirrup_spacing_mm(shear_kn: float, data: FrameInput) -> str:
    effective_depth = max(data.beam_depth_mm - 40.0, data.beam_depth_mm * 0.85)
    tau_v = shear_kn * 1000.0 / (data.beam_width_mm * effective_depth)
    tau_c_prelim = min(0.62, 0.16 * math.sqrt(data.concrete_grade_mpa))
    if tau_v <= tau_c_prelim:
        return "Nominal stirrups adequate preliminarily"
    return "Design shear reinforcement; start with closer spacing near supports"


def run_benchmark_checks(data: FrameInput) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []

    length = 4.0
    udl = 10.0
    test_element = Element(0, 0, 1, "beam", 1.0, 1.0, udl)
    fixed_end = fixed_end_load_local(test_element, length)
    expected_shear = udl * length / 2.0
    expected_moment = udl * length**2 / 12.0
    shear_error = abs(abs(fixed_end[1]) - expected_shear)
    moment_error = abs(abs(fixed_end[2]) - expected_moment)
    rows.append(
        {
            "Check": "Fixed-end beam under UDL",
            "Expected": f"V=wL/2={expected_shear:.2f} kN, M=wL2/12={expected_moment:.2f} kNm",
            "Observed": f"V={abs(fixed_end[1]):.2f} kN, M={abs(fixed_end[2]):.2f} kNm",
            "Status": "PASS" if shear_error < 1e-9 and moment_error < 1e-9 else "REVIEW",
        }
    )

    base_shear, storey_forces = calculate_lateral_loads(data)
    distribution_error = abs(sum(storey_forces) - base_shear)
    rows.append(
        {
            "Check": "Seismic storey-force distribution",
            "Expected": "Sum of storey forces equals base shear",
            "Observed": f"Base shear={base_shear:.3f} kN, Sum={sum(storey_forces):.3f} kN, Difference={distribution_error:.3e} kN",
            "Status": "PASS" if distribution_error < 1e-9 else "REVIEW",
        }
    )

    base_result = analyse_frame(data, SERVICE_COMBINATION, include_design_envelope=False)
    stiffer_data = replace(
        data,
        column_width_mm=data.column_width_mm * 1.4,
        column_depth_mm=data.column_depth_mm * 1.4,
    )
    stiffer_result = analyse_frame(stiffer_data, SERVICE_COMBINATION, include_design_envelope=False)
    base_sway = max(abs(base_result.displacements[3 * node.id]) for node in base_result.nodes) * 1000
    stiffer_sway = max(abs(stiffer_result.displacements[3 * node.id]) for node in stiffer_result.nodes) * 1000
    rows.append(
        {
            "Check": "Stiffness trend check",
            "Expected": "Increasing column size should reduce lateral displacement",
            "Observed": f"Original sway={base_sway:.3f} mm, stiffer-column sway={stiffer_sway:.3f} mm",
            "Status": "PASS" if stiffer_sway < base_sway else "REVIEW",
        }
    )

    return rows


def preliminary_checks(
    data: FrameInput,
    elements: List[Element],
    member_end_forces: Dict[int, np.ndarray],
    beam_diagrams: Dict[int, Dict[str, List[float]]],
    design_envelope: Dict[str, object] | None = None,
) -> Dict[int, Dict[str, str]]:
    checks: Dict[int, Dict[str, str]] = {}
    fck = data.concrete_grade_mpa
    fy = data.steel_grade_mpa
    beam_b = data.beam_width_mm
    beam_d_eff = max(data.beam_depth_mm - 40.0, data.beam_depth_mm * 0.85)
    col_b = data.column_width_mm
    col_d = data.column_depth_mm
    col_area_mm2 = col_b * col_d
    moment_factor = 0.138 if fy <= 415 else 0.133
    tau_c_prelim = min(0.62, 0.16 * math.sqrt(fck))
    member_envelope = (design_envelope or {}).get("members", {})

    beam_no = 0
    column_no = 0
    for element in elements:
        if element.kind == "beam":
            beam_no += 1
            diagram = beam_diagrams[element.id]
            envelope = member_envelope.get(element.id, {})
            mu = float(envelope.get("max_moment", diagram["max_moment"]))
            vu = float(envelope.get("max_shear", diagram["max_shear"]))
            moment_combo = envelope.get("moment_combo", "Service case")
            shear_combo = envelope.get("shear_combo", "Service case")
            mulim = moment_factor * fck * beam_b * beam_d_eff**2 / 1_000_000.0
            tau_v = vu * 1000.0 / (beam_b * beam_d_eff)
            ast_req = max(required_beam_steel_mm2(mu, data), minimum_beam_steel_mm2(data))
            bending = "OK" if mu <= mulim else "Resize / provide detailed reinforcement design"
            shear = "OK" if tau_v <= tau_c_prelim else "Needs shear reinforcement check"
            checks[element.id] = {
                "member": f"Beam {beam_no}",
                "bending": bending,
                "shear": shear,
                "details": (
                    f"Mu={mu:.2f} kNm ({moment_combo}), Mulim~{mulim:.2f} kNm, "
                    f"Ast_req~{ast_req:.0f} mm2, Vu={vu:.2f} kN ({shear_combo}), "
                    f"tau_v={tau_v:.2f} MPa, {nominal_stirrup_spacing_mm(vu, data)}"
                ),
            }
        else:
            column_no += 1
            forces = member_end_forces[element.id]
            envelope = member_envelope.get(element.id, {})
            axial = float(envelope.get("max_axial", max(abs(forces[0]), abs(forces[3]))))
            axial_combo = envelope.get("axial_combo", "Service case")
            pu_lim = 0.40 * fck * col_area_mm2 / 1000.0
            min_col_steel = 0.008 * col_area_mm2
            status = "OK" if axial <= pu_lim else "Increase column size / detailed interaction check needed"
            checks[element.id] = {
                "member": f"Column {column_no}",
                "axial": status,
                "details": (
                    f"Pu={axial:.2f} kN ({axial_combo}), preliminary axial capacity~{pu_lim:.2f} kN, "
                    f"minimum longitudinal steel~{min_col_steel:.0f} mm2"
                ),
            }
    return checks


def output_path(base_dir: Path, name: str) -> Path:
    base_dir.mkdir(parents=True, exist_ok=True)
    return base_dir / name


def create_run_folder(parent_dir: Path, data: FrameInput) -> Path:
    parent_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    label = (
        f"{stamp}_"
        f"{data.storeys}storey_"
        f"{data.bays}bay_"
        f"zone{data.seismic_zone}_"
        f"{data.support_condition}"
    )
    run_dir = parent_dir / label
    counter = 1
    while run_dir.exists():
        run_dir = parent_dir / f"{label}_{counter:02d}"
        counter += 1
    (run_dir / "plots").mkdir(parents=True)
    (run_dir / "reports").mkdir(parents=True)
    return run_dir


def unlocked_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def world_to_screen(
    x: float,
    y: float,
    xmin: float,
    xmax: float,
    ymin: float,
    ymax: float,
    width: int,
    height: int,
    margin: int,
) -> Tuple[int, int]:
    sx = margin + (x - xmin) / max(xmax - xmin, 1e-9) * (width - 2 * margin)
    sy = height - margin - (y - ymin) / max(ymax - ymin, 1e-9) * (height - 2 * margin)
    return int(sx), int(sy)


def draw_frame_plot(result: AnalysisResult, output_dir: Path) -> Tuple[Path, Path]:
    undeformed = output_path(output_dir, "undeformed_frame.png")
    deformed = output_path(output_dir, "deformed_shape.png")
    draw_structure(result, undeformed, deformed=False)
    draw_structure(result, deformed, deformed=True)
    return undeformed, deformed


def draw_structure(result: AnalysisResult, path: Path, deformed: bool) -> None:
    width, height, margin = 1100, 800, 80
    xs = [node.x for node in result.nodes]
    ys = [node.y for node in result.nodes]
    disp_scale = 1.0
    if deformed:
        max_dim = max(max(xs) - min(xs), max(ys) - min(ys), 1.0)
        max_disp = max(
            math.hypot(result.displacements[3 * n.id], result.displacements[3 * n.id + 1])
            for n in result.nodes
        )
        disp_scale = 0.08 * max_dim / max(max_disp, 1e-9)

    plotted_xs = list(xs)
    plotted_ys = list(ys)
    if deformed:
        plotted_xs.extend(node.x + result.displacements[3 * node.id] * disp_scale for node in result.nodes)
        plotted_ys.extend(node.y + result.displacements[3 * node.id + 1] * disp_scale for node in result.nodes)

    x_span = max(max(plotted_xs) - min(plotted_xs), result.inputs.bay_width_m, 1.0)
    y_span = max(max(plotted_ys) - min(plotted_ys), result.inputs.storey_height_m, 1.0)
    xmin = min(plotted_xs) - 0.12 * x_span
    xmax = max(plotted_xs) + 0.12 * x_span
    ymin = min(plotted_ys) - 0.12 * y_span
    ymax = max(plotted_ys) + 0.12 * y_span

    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    title = "Deformed shape" if deformed else "Undeformed RC frame"
    draw.text((margin, 25), title, fill=(30, 30, 30))
    draw.text((margin, 47), "Node numbers are shown beside joints. Displacements exaggerated for visibility.", fill=(90, 90, 90))

    def coord(node: Node) -> Tuple[int, int]:
        dx = result.displacements[3 * node.id] * disp_scale if deformed else 0.0
        dy = result.displacements[3 * node.id + 1] * disp_scale if deformed else 0.0
        return world_to_screen(node.x + dx, node.y + dy, xmin, xmax, ymin, ymax, width, height, margin)

    if deformed:
        for element in result.elements:
            p1 = world_to_screen(result.nodes[element.node_i].x, result.nodes[element.node_i].y, xmin, xmax, ymin, ymax, width, height, margin)
            p2 = world_to_screen(result.nodes[element.node_j].x, result.nodes[element.node_j].y, xmin, xmax, ymin, ymax, width, height, margin)
            draw.line([p1, p2], fill=(190, 190, 190), width=2)

    for element in result.elements:
        p1 = coord(result.nodes[element.node_i])
        p2 = coord(result.nodes[element.node_j])
        color = (34, 92, 160) if element.kind == "beam" else (38, 125, 87)
        draw.line([p1, p2], fill=color, width=5)
    for node in result.nodes:
        x, y = coord(node)
        draw.ellipse((x - 5, y - 5, x + 5, y + 5), fill=(20, 20, 20))
        label_x = x + 8
        label_y = y - 16
        draw.rectangle((label_x - 2, label_y - 2, label_x + 22, label_y + 12), fill=(255, 255, 255))
        draw.text((label_x, label_y), str(node.id), fill=(25, 25, 25))
    image.save(path)


def grid_label(grid_index: int) -> str:
    label = ""
    index = grid_index
    while True:
        label = chr(ord("A") + index % 26) + label
        index = index // 26 - 1
        if index < 0:
            return label


def node_location(result: AnalysisResult, node_id: int) -> Tuple[int, str]:
    level = node_id // (result.inputs.bays + 1)
    grid = node_id % (result.inputs.bays + 1)
    return level, grid_label(grid)


def member_display_numbers(elements: List[Element]) -> Dict[int, int]:
    numbers: Dict[int, int] = {}
    counters = {"beam": 0, "column": 0}
    for element in elements:
        counters[element.kind] += 1
        numbers[element.id] = counters[element.kind]
    return numbers


def draw_beam_diagrams(result: AnalysisResult, output_dir: Path) -> List[Path]:
    beams = [element for element in result.elements if element.kind == "beam"]
    images: List[Path] = []
    beams_per_page = 5
    for key, title, stem in [
        ("shear", "Beam Shear Force Diagrams (kN)", "beam_sfd"),
        ("moment", "Beam Bending Moment Diagrams (kNm)", "beam_bmd"),
    ]:
        pages = max(1, math.ceil(len(beams) / beams_per_page))
        for page_index in range(pages):
            selected = beams[page_index * beams_per_page : (page_index + 1) * beams_per_page]
            path = output_path(output_dir, f"{stem}_page_{page_index + 1:02d}.png")
            draw_diagram_grid(result, selected, path, key, title, page_index + 1, pages)
            images.append(path)
    return images


def draw_diagram_grid(
    result: AnalysisResult,
    beams: List[Element],
    path: Path,
    key: str,
    title: str,
    page_number: int,
    total_pages: int,
) -> None:
    width, row_height, margin = 1200, 145, 70
    height = max(420, margin + 80 + row_height * len(beams))
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    color = (196, 64, 54) if key == "moment" else (31, 119, 180)
    unit = "kNm" if key == "moment" else "kN"
    display_numbers = member_display_numbers(result.elements)
    draw.text((margin, 24), f"{title} - Page {page_number} of {total_pages}", fill=(30, 30, 30))
    draw.text(
        (margin, 48),
        "Each row is one beam. The grey line is the beam axis; values above/below it show sign.",
        fill=(80, 80, 80),
    )

    left, right = 220, width - 120
    diagram_width = right - left
    for index, beam in enumerate(beams):
        level = beam.node_i // (result.inputs.bays + 1)
        bay = beam.node_i % (result.inputs.bays + 1) + 1
        diagram = result.beam_diagrams[beam.id]
        values = diagram[key]
        xs = diagram["x"]
        max_value = max(abs(v) for v in values) or 1.0
        y0 = margin + 55 + index * row_height
        axis_y = y0 + row_height // 2
        scale = row_height * 0.34 / max_value

        draw.text((margin, y0 + 8), f"Beam {display_numbers[beam.id]}", fill=(25, 25, 25))
        draw.text((margin, y0 + 30), f"Storey {level}, Bay {bay}", fill=(80, 80, 80))
        draw.text((margin, y0 + 52), f"Max |{unit}| = {max_value:.2f}", fill=(80, 80, 80))
        draw.line([(left, axis_y), (right, axis_y)], fill=(145, 145, 145), width=2)
        draw.text((left - 10, axis_y + 8), "i", fill=(80, 80, 80))
        draw.text((right - 5, axis_y + 8), "j", fill=(80, 80, 80))
        draw.text((left, axis_y - int(max_value * scale) - 18), f"+{max_value:.1f}", fill=(100, 100, 100))
        draw.text((left, axis_y + int(max_value * scale) + 6), f"-{max_value:.1f}", fill=(100, 100, 100))

        points = []
        for x, value in zip(xs, values):
            px = left + x / max(xs[-1], 1e-9) * diagram_width
            py = axis_y - value * scale
            points.append((int(px), int(py)))
        if len(points) > 1:
            draw.line(points, fill=color, width=3)
            for point in points[:: max(1, len(points) // 6)]:
                draw.line([(point[0], axis_y), point], fill=(220, 220, 220), width=1)
        draw.text((left, axis_y - 28), f"{values[0]:.2f} {unit}", fill=color)
        draw.text((right - 95, axis_y - 28), f"{values[-1]:.2f} {unit}", fill=color)
        draw.line([(margin, y0 + row_height - 6), (width - margin, y0 + row_height - 6)], fill=(235, 235, 235), width=1)
    image.save(path)


def make_table_rows(result: AnalysisResult) -> Dict[str, List[List[object]]]:
    nodes = []
    for node in result.nodes:
        storey, grid = node_location(result, node.id)
        nodes.append(
            [
                node.id,
                storey,
                grid,
                node.x,
                node.y,
                result.displacements[3 * node.id] * 1000,
                result.displacements[3 * node.id + 1] * 1000,
                result.displacements[3 * node.id + 2],
            ]
        )

    reactions = []
    for node in result.nodes[: result.inputs.bays + 1]:
        storey, grid = node_location(result, node.id)
        reactions.append(
            [
                node.id,
                storey,
                grid,
                result.reactions[3 * node.id],
                result.reactions[3 * node.id + 1],
                result.reactions[3 * node.id + 2],
            ]
        )

    members = []
    display_numbers = member_display_numbers(result.elements)
    for element in result.elements:
        forces = result.member_end_forces[element.id]
        check = result.checks[element.id]
        member_type = "Beam" if element.kind == "beam" else "Column"
        members.append(
            [
                f"{member_type} {display_numbers[element.id]}",
                element.node_i,
                element.node_j,
                forces[0],
                forces[1],
                forces[2],
                forces[3],
                forces[4],
                forces[5],
                check["details"],
            ]
        )
    verification = [[name, value, note] for name, (value, note) in result.verification.items()]
    combinations = [
        [
            row["Combination"],
            row["Max sway (mm)"],
            row["Max beam moment (kNm)"],
            row["Max beam shear (kN)"],
            row["Max column axial (kN)"],
        ]
        for row in result.design_envelope.get("combinations", [])
    ]
    benchmarks = [
        [row["Check"], row["Expected"], row["Observed"], row["Status"]]
        for row in result.benchmark_results
    ]
    return {
        "nodes": nodes,
        "reactions": reactions,
        "members": members,
        "verification": verification,
        "combinations": combinations,
        "benchmarks": benchmarks,
    }


def write_excel_report(result: AnalysisResult, image_paths: Iterable[Path], output_dir: Path) -> Path:
    path = unlocked_output_path(output_path(output_dir, "rc_frame_analysis_report.xlsx"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    title_fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"] = "Automated 2D RC Frame Analysis & Preliminary Design"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    ws.append([])
    for label, value in input_summary(result.inputs):
        ws.append([label, value])
    ws.append(["Base shear (kN)", result.base_shear_kn])
    ws.append(["Service load case", result.load_combination.name])
    summary = result.design_envelope.get("summary", {})
    if summary:
        ws.append(["Governing beam moment", f"{summary['max_beam_moment']:.2f} kNm | {summary['max_beam_moment_combo']}"])
        ws.append(["Governing beam shear", f"{summary['max_beam_shear']:.2f} kN | {summary['max_beam_shear_combo']}"])
        ws.append(["Governing column axial", f"{summary['max_column_axial']:.2f} kN | {summary['max_column_axial_combo']}"])
    for i, force in enumerate(result.storey_forces_kn, start=1):
        ws.append([f"Storey {i} lateral force (kN)", force])
    for column in range(1, 4):
        ws.column_dimensions[get_column_letter(column)].width = 28

    tables = make_table_rows(result)
    add_sheet(wb, "Node Map + Displacements", ["Node", "Storey", "Grid", "X m", "Y m", "Ux mm", "Uy mm", "Rz rad"], tables["nodes"])
    add_sheet(wb, "Support Reactions", ["Node", "Storey", "Grid", "Rx kN", "Ry kN", "Mz kNm"], tables["reactions"])
    add_sheet(
        wb,
        "Member Forces",
        ["Member", "Node i", "Node j", "Ni kN", "Vi kN", "Mi kNm", "Nj kN", "Vj kN", "Mj kNm", "Check"],
        tables["members"],
    )
    add_sheet(wb, "Verification", ["Check", "Value", "Expected"], tables["verification"])
    add_sheet(
        wb,
        "Load Combinations",
        ["Combination", "Max sway mm", "Max beam moment kNm", "Max beam shear kN", "Max column axial kN"],
        tables["combinations"],
    )
    add_sheet(wb, "Benchmarks", ["Check", "Expected", "Observed", "Status"], tables["benchmarks"])
    img_sheet = wb.create_sheet("Plots")
    row = 1
    for image_path in image_paths:
        img_sheet.cell(row=row, column=1, value=image_path.stem.replace("_", " ").title())
        img = XlsxImage(str(image_path))
        img.width = 520
        img.height = 380
        img_sheet.add_image(img, f"A{row + 1}")
        row += 22
    wb.save(path)
    return path


def add_sheet(wb: Workbook, name: str, headers: List[str], rows: List[List[object]]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([round(v, 4) if isinstance(v, float) else v for v in row])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    if name == "Verification":
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 90
        ws.column_dimensions["C"].width = 70


def input_summary(data: FrameInput) -> List[Tuple[str, object]]:
    return [
        ("Storeys", data.storeys),
        ("Bays", data.bays),
        ("Bay width (m)", data.bay_width_m),
        ("Storey height (m)", data.storey_height_m),
        ("Beam size (mm)", f"{data.beam_width_mm:g} x {data.beam_depth_mm:g}"),
        ("Column size (mm)", f"{data.column_width_mm:g} x {data.column_depth_mm:g}"),
        ("Concrete grade", f"M{data.concrete_grade_mpa:g}"),
        ("Steel grade", f"Fe{data.steel_grade_mpa:g}"),
        ("Dead load (kN/m)", data.dead_load_kn_m),
        ("Live load (kN/m)", data.live_load_kn_m),
        ("Seismic zone", data.seismic_zone),
        ("Support condition", data.support_condition),
    ]


def write_pdf_report(result: AnalysisResult, image_paths: Iterable[Path], output_dir: Path) -> Path:
    path = unlocked_output_path(output_path(output_dir, "rc_frame_analysis_report.pdf"))
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=14 * mm, leftMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Automated 2D RC Frame Analysis & Preliminary Design", styles["Title"]),
        Spacer(1, 8),
        Paragraph(
            "This report presents a 2D RC frame analysis using the matrix stiffness method. "
            "Preliminary checks are based on simplified IS 456 and IS 1893 concepts and are not a substitute for detailed design.",
            styles["BodyText"],
        ),
        Spacer(1, 10),
        Paragraph("Input Summary", styles["Heading2"]),
        simple_pdf_table([["Parameter", "Value"]] + [[a, str(b)] for a, b in input_summary(result.inputs)]),
        Spacer(1, 8),
        Paragraph(f"Calculated base shear: {result.base_shear_kn:.2f} kN", styles["BodyText"]),
        Paragraph(f"Plots are based on service load case: {result.load_combination.name}.", styles["BodyText"]),
        Paragraph("IS 1893 zone factors used: II=0.10, III=0.16, IV=0.24, V=0.36.", styles["BodyText"]),
        Paragraph(
            "Strength design checks use an envelope of IS-style load combinations including "
            "1.5(DL+LL), 1.2(DL+LL+/-EQ), 1.5(DL+/-EQ), and 0.9DL+/-1.5EQ.",
            styles["BodyText"],
        ),
        Paragraph(
            "SFD/BMD sign convention: each beam is plotted from left node i to right node j. "
            "Values above the beam axis are positive and values below the axis are negative.",
            styles["BodyText"],
        ),
        Paragraph(
            "Node numbering starts at the bottom-left support and moves left-to-right along each floor, "
            "then continues upward storey by storey. Grid A is the leftmost column line.",
            styles["BodyText"],
        ),
    ]
    for image_path in image_paths:
        story.append(PageBreak())
        story.append(Paragraph(image_path.stem.replace("_", " ").title(), styles["Heading2"]))
        if "beam_sfd_page" in image_path.stem or "beam_bmd_page" in image_path.stem:
            story.append(PdfImage(str(image_path), width=255 * mm, height=170 * mm))
        else:
            story.append(PdfImage(str(image_path), width=230 * mm, height=167 * mm))

    story.append(PageBreak())
    story.append(Paragraph("Key Results", styles["Heading2"]))
    tables = make_table_rows(result)
    summary = result.design_envelope.get("summary", {})
    if summary:
        story.append(Paragraph("Design Envelope Summary", styles["Heading3"]))
        story.append(
            simple_pdf_table(
                [
                    ["Demand", "Governing value", "Combination"],
                    ["Beam bending moment", f"{summary['max_beam_moment']:.2f} kNm", summary["max_beam_moment_combo"]],
                    ["Beam shear", f"{summary['max_beam_shear']:.2f} kN", summary["max_beam_shear_combo"]],
                    ["Column axial force", f"{summary['max_column_axial']:.2f} kN", summary["max_column_axial_combo"]],
                    ["Lateral displacement", f"{summary['max_lateral_displacement']:.3f} mm", summary["max_lateral_displacement_combo"]],
                ],
                font_size=7,
                col_widths=[64 * mm, 62 * mm, 120 * mm],
            )
        )
        story.append(Spacer(1, 8))
    story.append(Paragraph("Load Combination Results", styles["Heading3"]))
    story.append(simple_pdf_table([["Combination", "Max sway mm", "Max beam moment kNm", "Max beam shear kN", "Max column axial kN"]] + tables["combinations"], font_size=6, col_widths=[82 * mm, 38 * mm, 45 * mm, 42 * mm, 42 * mm]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Benchmark Validation", styles["Heading3"]))
    story.append(simple_pdf_table([["Check", "Expected", "Observed", "Status"]] + tables["benchmarks"], font_size=6, col_widths=[58 * mm, 70 * mm, 96 * mm, 24 * mm]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Node Map and Nodal Displacements", styles["Heading3"]))
    story.append(simple_pdf_table([["Node", "Storey", "Grid", "X m", "Y m", "Ux mm", "Uy mm", "Rz rad"]] + [[r[0], r[1], r[2], f"{r[3]:.2f}", f"{r[4]:.2f}", f"{r[5]:.3f}", f"{r[6]:.3f}", f"{r[7]:.6f}"] for r in tables["nodes"][:25]], font_size=7))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Support Reactions", styles["Heading3"]))
    story.append(simple_pdf_table([["Node", "Storey", "Grid", "Rx kN", "Ry kN", "Mz kNm"]] + [[r[0], r[1], r[2], f"{r[3]:.2f}", f"{r[4]:.2f}", f"{r[5]:.2f}"] for r in tables["reactions"]], font_size=7))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Verification Checks", styles["Heading3"]))
    story.append(simple_pdf_table([["Check", "Value", "Expected"]] + [[r[0], r[1], r[2]] for r in tables["verification"]], font_size=6, col_widths=[48 * mm, 132 * mm, 80 * mm]))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Member Checks", styles["Heading3"]))
    check_rows = [["Member", "Preliminary check", "Details"]]
    for check in result.checks.values():
        status = check.get("bending") or check.get("axial") or "Review"
        if check.get("shear"):
            status = f"{status}; shear: {check['shear']}"
        check_rows.append([check["member"], status, check["details"]])
    story.append(simple_pdf_table(check_rows, font_size=6, col_widths=[28 * mm, 64 * mm, 168 * mm]))
    doc.build(story)
    return path


def write_run_summary(result: AnalysisResult, run_dir: Path, pdf_path: Path, excel_path: Path, image_paths: Iterable[Path]) -> Path:
    path = run_dir / "run_summary.txt"
    max_sway_mm = max(abs(result.displacements[3 * node.id]) for node in result.nodes) * 1000
    max_beam_moment = max(d["max_moment"] for d in result.beam_diagrams.values())
    max_beam_shear = max(d["max_shear"] for d in result.beam_diagrams.values())
    lines = [
        "RC Frame Analysis Run Summary",
        "=" * 30,
        f"Created: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "Model",
        f"- Storeys: {result.inputs.storeys}",
        f"- Bays: {result.inputs.bays}",
        f"- Bay width: {result.inputs.bay_width_m} m",
        f"- Storey height: {result.inputs.storey_height_m} m",
        f"- Beam: {result.inputs.beam_width_mm:g} x {result.inputs.beam_depth_mm:g} mm",
        f"- Column: {result.inputs.column_width_mm:g} x {result.inputs.column_depth_mm:g} mm",
        f"- Concrete: M{result.inputs.concrete_grade_mpa:g}",
        f"- Steel: Fe{result.inputs.steel_grade_mpa:g}",
        f"- Seismic zone: {result.inputs.seismic_zone}",
        f"- Support: {result.inputs.support_condition}",
        "",
        "Node Numbering",
        "- Node numbering starts at the bottom-left support.",
        "- Numbers increase left-to-right along each level.",
        "- Numbering then continues upward storey by storey.",
        "- Grid A is the leftmost column line.",
        "",
        "Key Results",
        f"- Base shear: {result.base_shear_kn:.2f} kN",
        f"- Maximum lateral displacement: {max_sway_mm:.3f} mm",
        f"- Maximum beam moment: {max_beam_moment:.2f} kNm",
        f"- Maximum beam shear: {max_beam_shear:.2f} kN",
    ]
    summary = result.design_envelope.get("summary", {})
    if summary:
        lines.extend(
            [
                "",
                "Design Envelope",
                f"- Max beam moment: {summary['max_beam_moment']:.2f} kNm ({summary['max_beam_moment_combo']})",
                f"- Max beam shear: {summary['max_beam_shear']:.2f} kN ({summary['max_beam_shear_combo']})",
                f"- Max column axial: {summary['max_column_axial']:.2f} kN ({summary['max_column_axial_combo']})",
                f"- Max lateral displacement: {summary['max_lateral_displacement']:.3f} mm ({summary['max_lateral_displacement_combo']})",
            ]
        )
    if result.benchmark_results:
        lines.append("")
        lines.append("Benchmark Validation")
        for row in result.benchmark_results:
            lines.append(f"- {row['Check']}: {row['Status']} | {row['Observed']}")
    lines.extend(
        [
            "",
            "Files",
            f"- PDF report: {pdf_path.name}",
            f"- Excel report: {excel_path.name}",
        ]
    )
    for image_path in image_paths:
        lines.append(f"- Plot: plots/{image_path.name}")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def simple_pdf_table(rows: List[List[object]], font_size: int = 8, col_widths: List[float] | None = None) -> Table:
    styles = getSampleStyleSheet()
    cell_style = styles["BodyText"]
    cell_style.fontSize = font_size
    cell_style.leading = font_size + 2
    wrapped_rows = [[Paragraph(str(cell), cell_style) for cell in row] for row in rows]
    table = Table(wrapped_rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def print_console_summary(result: AnalysisResult, pdf_path: Path, excel_path: Path, image_paths: Iterable[Path]) -> None:
    max_sway_mm = max(abs(result.displacements[3 * node.id]) for node in result.nodes) * 1000
    max_beam_moment = max(d["max_moment"] for d in result.beam_diagrams.values())
    max_beam_shear = max(d["max_shear"] for d in result.beam_diagrams.values())
    print("\nAnalysis complete.")
    print(f"Nodes: {len(result.nodes)} | Elements: {len(result.elements)}")
    print(f"Base shear: {result.base_shear_kn:.2f} kN")
    print(f"Maximum lateral displacement: {max_sway_mm:.3f} mm")
    print(f"Maximum beam moment: {max_beam_moment:.2f} kNm")
    print(f"Maximum beam shear: {max_beam_shear:.2f} kN")
    summary = result.design_envelope.get("summary", {})
    if summary:
        print("Design envelope:")
        print(f"- Beam moment: {summary['max_beam_moment']:.2f} kNm ({summary['max_beam_moment_combo']})")
        print(f"- Beam shear: {summary['max_beam_shear']:.2f} kN ({summary['max_beam_shear_combo']})")
        print(f"- Column axial: {summary['max_column_axial']:.2f} kN ({summary['max_column_axial_combo']})")
    if result.benchmark_results:
        print("Benchmark checks:")
        for row in result.benchmark_results:
            print(f"- {row['Check']}: {row['Status']}")
    print("Verification residuals:")
    for name, (value, note) in result.verification.items():
        print(f"- {name}: {value} ({note})")
    print("\nGenerated files:")
    print(f"- PDF report: {pdf_path}")
    print(f"- Excel report: {excel_path}")
    for path in image_paths:
        print(f"- Plot: {path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Analyse and preliminarily design a 2D RC building frame.")
    parser.add_argument("--demo", action="store_true", help="Run with built-in sample inputs.")
    parser.add_argument("--gui", action="store_true", help="Open the desktop input form.")
    parser.add_argument("--output", default="outputs", help="Output folder for plots and reports.")
    return parser.parse_args()


def launch_gui() -> None:
    defaults = FrameInput()
    root = tk.Tk()
    root.title("Automated 2D RC Frame Analysis & Preliminary Design")
    root.geometry("760x640")
    root.minsize(720, 600)

    main = ttk.Frame(root, padding=16)
    main.pack(fill="both", expand=True)
    title = ttk.Label(main, text="Automated 2D RC Frame Analysis & Preliminary Design", font=("Segoe UI", 16, "bold"))
    title.grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 12))

    fields = [
        ("storeys", "Number of storeys", defaults.storeys),
        ("bays", "Number of bays", defaults.bays),
        ("bay_width_m", "Bay width (m)", defaults.bay_width_m),
        ("storey_height_m", "Storey height (m)", defaults.storey_height_m),
        ("beam_width_mm", "Beam width (mm)", defaults.beam_width_mm),
        ("beam_depth_mm", "Beam depth (mm)", defaults.beam_depth_mm),
        ("column_width_mm", "Column width (mm)", defaults.column_width_mm),
        ("column_depth_mm", "Column depth (mm)", defaults.column_depth_mm),
        ("concrete_grade_mpa", "Concrete grade fck (MPa)", defaults.concrete_grade_mpa),
        ("steel_grade_mpa", "Steel grade fy (MPa)", defaults.steel_grade_mpa),
        ("dead_load_kn_m", "Dead load (kN/m)", defaults.dead_load_kn_m),
        ("live_load_kn_m", "Live load (kN/m)", defaults.live_load_kn_m),
        ("lateral_load_kn", "Manual total lateral load (kN)", defaults.lateral_load_kn),
        ("importance_factor", "Importance factor", defaults.importance_factor),
        ("response_reduction_factor", "Response reduction factor", defaults.response_reduction_factor),
    ]
    entries: Dict[str, tk.StringVar] = {}
    for index, (key, label, value) in enumerate(fields):
        row = index // 2 + 1
        col = (index % 2) * 2
        ttk.Label(main, text=label).grid(row=row, column=col, sticky="w", padx=(0, 8), pady=5)
        var = tk.StringVar(value=str(value))
        entries[key] = var
        ttk.Entry(main, textvariable=var, width=18).grid(row=row, column=col + 1, sticky="ew", pady=5)

    zone_var = tk.StringVar(value=defaults.seismic_zone)
    support_var = tk.StringVar(value=defaults.support_condition)
    output_var = tk.StringVar(value=str(Path.cwd() / "outputs"))
    option_row = len(fields) // 2 + 2

    ttk.Label(main, text="Seismic zone").grid(row=option_row, column=0, sticky="w", padx=(0, 8), pady=5)
    ttk.Combobox(main, textvariable=zone_var, values=["II", "III", "IV", "V"], width=15, state="readonly").grid(
        row=option_row, column=1, sticky="ew", pady=5
    )
    ttk.Label(main, text="Support condition").grid(row=option_row, column=2, sticky="w", padx=(16, 8), pady=5)
    ttk.Combobox(main, textvariable=support_var, values=["fixed", "pinned", "roller"], width=15, state="readonly").grid(
        row=option_row, column=3, sticky="ew", pady=5
    )

    ttk.Label(main, text="Output folder").grid(row=option_row + 1, column=0, sticky="w", padx=(0, 8), pady=8)
    ttk.Entry(main, textvariable=output_var).grid(row=option_row + 1, column=1, columnspan=2, sticky="ew", pady=8)

    def browse_output() -> None:
        selected = filedialog.askdirectory(initialdir=output_var.get() or str(Path.cwd()))
        if selected:
            output_var.set(selected)

    ttk.Button(main, text="Browse", command=browse_output).grid(row=option_row + 1, column=3, sticky="ew", pady=8)

    status_var = tk.StringVar(value="Enter values and click Run Analysis.")
    status = ttk.Label(main, textvariable=status_var, foreground="#1F4E78")
    status.grid(row=option_row + 3, column=0, columnspan=4, sticky="w", pady=(8, 0))

    result_box = tk.Text(main, height=8, wrap="word")
    result_box.grid(row=option_row + 4, column=0, columnspan=4, sticky="nsew", pady=(8, 0))
    result_box.insert("1.0", "Results will appear here after analysis.\n")
    result_box.configure(state="disabled")

    def build_input_from_form() -> FrameInput:
        data = FrameInput(
            storeys=int(entries["storeys"].get()),
            bays=int(entries["bays"].get()),
            bay_width_m=float(entries["bay_width_m"].get()),
            storey_height_m=float(entries["storey_height_m"].get()),
            beam_width_mm=float(entries["beam_width_mm"].get()),
            beam_depth_mm=float(entries["beam_depth_mm"].get()),
            column_width_mm=float(entries["column_width_mm"].get()),
            column_depth_mm=float(entries["column_depth_mm"].get()),
            concrete_grade_mpa=float(entries["concrete_grade_mpa"].get()),
            steel_grade_mpa=float(entries["steel_grade_mpa"].get()),
            dead_load_kn_m=float(entries["dead_load_kn_m"].get()),
            live_load_kn_m=float(entries["live_load_kn_m"].get()),
            seismic_zone=zone_var.get(),
            lateral_load_kn=float(entries["lateral_load_kn"].get()),
            importance_factor=float(entries["importance_factor"].get()),
            response_reduction_factor=float(entries["response_reduction_factor"].get()),
            support_condition=support_var.get(),
        )
        validate_inputs(data)
        return data

    def run_from_form() -> None:
        try:
            status_var.set("Running analysis and generating reports...")
            root.update_idletasks()
            data = build_input_from_form()
            result = analyse_frame(data)
            run_dir = create_run_folder(Path(output_var.get()), data)
            plots_dir = run_dir / "plots"
            reports_dir = run_dir / "reports"
            frame_images = draw_frame_plot(result, plots_dir)
            diagram_images = draw_beam_diagrams(result, plots_dir)
            image_paths = [*frame_images, *diagram_images]
            excel_path = write_excel_report(result, image_paths, reports_dir)
            pdf_path = write_pdf_report(result, image_paths, reports_dir)
            summary_path = write_run_summary(result, run_dir, pdf_path, excel_path, image_paths)
            max_sway_mm = max(abs(result.displacements[3 * node.id]) for node in result.nodes) * 1000
            max_beam_moment = max(d["max_moment"] for d in result.beam_diagrams.values())
            max_beam_shear = max(d["max_shear"] for d in result.beam_diagrams.values())
            envelope_summary = result.design_envelope.get("summary", {})
            envelope_text = ""
            if envelope_summary:
                envelope_text = (
                    f"\nDesign envelope:\n"
                    f"Beam moment: {envelope_summary['max_beam_moment']:.2f} kNm "
                    f"({envelope_summary['max_beam_moment_combo']})\n"
                    f"Beam shear: {envelope_summary['max_beam_shear']:.2f} kN "
                    f"({envelope_summary['max_beam_shear_combo']})\n"
                    f"Column axial: {envelope_summary['max_column_axial']:.2f} kN "
                    f"({envelope_summary['max_column_axial_combo']})\n"
                )
            summary = (
                f"Analysis complete.\n\n"
                f"Nodes: {len(result.nodes)}\n"
                f"Elements: {len(result.elements)}\n"
                f"Base shear: {result.base_shear_kn:.2f} kN\n"
                f"Maximum lateral displacement: {max_sway_mm:.3f} mm\n"
                f"Maximum beam moment: {max_beam_moment:.2f} kNm\n"
                f"Maximum beam shear: {max_beam_shear:.2f} kN\n"
                f"{envelope_text}\n"
                f"PDF report:\n{pdf_path}\n\n"
                f"Excel report:\n{excel_path}\n\n"
                f"Run summary:\n{summary_path}\n\n"
                f"Run folder:\n{run_dir}"
            )
            result_box.configure(state="normal")
            result_box.delete("1.0", "end")
            result_box.insert("1.0", summary)
            result_box.configure(state="disabled")
            status_var.set("Done. Reports and plots have been created.")
            messagebox.showinfo("Analysis complete", "Reports and plots have been created successfully.")
        except Exception as exc:
            status_var.set("Please check the input values.")
            messagebox.showerror("Could not run analysis", str(exc))

    ttk.Button(main, text="Run Analysis", command=run_from_form).grid(
        row=option_row + 2, column=0, columnspan=4, sticky="ew", pady=(12, 4)
    )
    for col in range(4):
        main.columnconfigure(col, weight=1)
    main.rowconfigure(option_row + 4, weight=1)
    root.mainloop()


def main() -> None:
    ensure_dependencies()
    args = parse_args()
    if args.gui:
        launch_gui()
        return
    data = FrameInput() if args.demo else read_inputs_interactively()
    result = analyse_frame(data)
    run_dir = create_run_folder(Path(args.output), data)
    plots_dir = run_dir / "plots"
    reports_dir = run_dir / "reports"
    frame_images = draw_frame_plot(result, plots_dir)
    diagram_images = draw_beam_diagrams(result, plots_dir)
    image_paths = [*frame_images, *diagram_images]
    excel_path = write_excel_report(result, image_paths, reports_dir)
    pdf_path = write_pdf_report(result, image_paths, reports_dir)
    summary_path = write_run_summary(result, run_dir, pdf_path, excel_path, image_paths)
    print_console_summary(result, pdf_path, excel_path, image_paths)
    print(f"- Run summary: {summary_path}")
    print(f"- Run folder: {run_dir}")


if __name__ == "__main__":
    main()
